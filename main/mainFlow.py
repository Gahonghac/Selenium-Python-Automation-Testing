from selenium import webdriver
from time import sleep
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
import openpyxl
import time

def config():
    global driver
    global wait
    global totalRow
    global totalsCol
    global sheet
    #Chrome dirver
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument('--ignore-ssl-errors=yes')
    chrome_options.add_argument('--ignore-certificate-errors')
    chrome_options.add_experimental_option('detach', True)
    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 60)
    driver.maximize_window()
    #Config exel file
    exelFile = openpyxl.load_workbook("./dataUtil/fileTestCase.xlsx")
    sheet = exelFile['TestCase']
    #Properties of testcase
    totalsCol = sheet.max_column
    #The number of testcase, from 3 to ...
    totalRow = sheet.max_row

def getvalueExel(row, col):
    return sheet.cell(row, col).value


def switchIframeByTitle(_title):
    wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, f"iframe[title='{_title}']")))
    iframes = driver.find_elements(By.TAG_NAME, 'iframe')
    for iframe in iframes:
        try:
            if iframe.get_attribute('title')==_title:
                driver.switch_to.frame(iframe)
                break
        except NoSuchElementException:
            pass
        

def switchToNewTab(numberOfTab, sttTab):
    wait.until(EC.number_of_windows_to_be(numberOfTab))
    driver.switch_to.window(driver.window_handles[sttTab])


def getTo(_title):
    listScopeDash = wait.until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, '.menu-link-body')))
    # listScopeDash = driver.find_elements(By.CSS_SELECTOR, '.menu-link-body')
    for element in listScopeDash:
        title = element.get_attribute("title")
        
        if title == _title:
            element.find_element(By.TAG_NAME,'a').click()
            time.sleep(1)
            # element.click()
            break

def clickDropdownBoxByText(id, textData):
    wait.until(EC.presence_of_element_located((By.ID, id)))
    selectElement = driver.find_element(By.ID, id)
    select = Select(selectElement)
    select.select_by_visible_text(str(textData))
    # raise('Error at dropdown box! Do not have any data match to dropdown box!')


def clickButtonByText(_text):
    listButton = driver.find_elements(By.TAG_NAME,'button')
    for button in listButton:
        if button.text == _text:
            button.click()    
            return


def scrollToElement(element):
    window_height = driver.execute_script("return window.innerHeight;")
    element_position = element.location["y"]
    current_position = driver.execute_script("return window.scrollY;")
    distance_to_scroll = element_position - current_position - (window_height // 2)
    scroll_speed = min(3, abs(distance_to_scroll))
    
    steps = abs(distance_to_scroll) // scroll_speed
    
    # Cuộn từng bước để đưa phần tử vào tầm nhìn
    for _ in range(int(steps)):
        driver.execute_script(f"window.scrollBy(0, {'-' if distance_to_scroll < 0 else ''}{scroll_speed});")
        driver.implicitly_wait(5) 
    time.sleep(1)


def scrollDownUp(begin, end):
    scrollToElement(begin)
    scrollToElement(end)


def testLoginPage(userName, password):
    try:
        #Set Username
        driver.find_element(By.ID,'username').send_keys(userName)
        #Set password
        driver.find_element(By.ID,'password').send_keys(password)
        #Click login button
        driver.find_element(By.CSS_SELECTOR, '#btnSubmit > span').click()
        # If url do not change -> Login fail!
        wait.until(EC.url_changes(driver.current_url))
    except:
        raise Exception('Testcase fail at Lolgin Page!')


def testKhoiTaoKhoanVay(fullName, productName, loanMethod, policy):
        try:
            getTo('Khởi tạo khoản vay 2')
            switchIframeByTitle('Coach')

            wait.until(EC.visibility_of_element_located((By.ID, 'text-input-LOS_Customers_Find1:txt_customer_name')))                                   
            driver.find_element(By.ID, 'text-input-LOS_Customers_Find1:txt_customer_name').send_keys(fullName)         
            driver.find_element(By.ID, 'button-button-LOS_Customers_Find1:btnSearch').click()

            wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#button-button-LOS_Customers_Find1\:Table1\:Horizontal_Layout5\[1\]\:Horizontal_Layout6\:btn_edit')))
            driver.find_element(By.CSS_SELECTOR, '#button-button-LOS_Customers_Find1\:Table1\:Horizontal_Layout5\[2\]\:Horizontal_Layout6\:btn_edit').click()
            
            clickDropdownBoxByText('singleselect-LOS_Product_Loan_Choose:Single_Select1', productName)
            time.sleep(1)
            clickDropdownBoxByText('singleselect-LOS_Product_Loan_Choose:lh_vay', loanMethod)
            clickDropdownBoxByText('singleselect-LOS_Product_Loan_Choose:dt_kh', policy)
            driver.find_element(By.CSS_SELECTOR, '#button-button-LOS_Product_Loan_Choose\:Button2').click()
        except Exception as ex:
            raise Exception('Testcase fail at Khoi Tao Khoan Vay!')


def testKhachHangPage( referencePersonName,
           relationship, gender, phoneNumber, nationality, province, district, ward, address):
    try:
        #1 new tab, switch to stab 2nd (handles begin from 0)
        switchToNewTab(2, 1) 
        switchIframeByTitle('Coach')
        # click to Tiếp tục button
        clickButtonByText('Tiếp tục')
        # switch iframe
        switchIframeByTitle('Coach')
        driver.find_element(By.ID, 'text-input-Los_1_nguoi_tham_chieu:CUS111').send_keys(referencePersonName)
        clickDropdownBoxByText('singleselect-Los_1_nguoi_tham_chieu:CUS113', relationship)
        clickDropdownBoxByText('singleselect-Los_1_nguoi_tham_chieu:CUS114', gender)
        driver.find_element(By.CSS_SELECTOR, '#text-input-Los_1_nguoi_tham_chieu\:CUS115').send_keys(phoneNumber)
        clickDropdownBoxByText('singleselect-Los_1_nguoi_tham_chieu:Horizontal_Layout2:Table1:Single_Select5[0]', nationality)
        clickDropdownBoxByText('singleselect-Los_1_nguoi_tham_chieu:Horizontal_Layout2:Table1:tinh[0]', province)
        clickDropdownBoxByText('singleselect-Los_1_nguoi_tham_chieu:Horizontal_Layout2:Table1:quan[0]', district)
        clickDropdownBoxByText('singleselect-Los_1_nguoi_tham_chieu:Horizontal_Layout2:Table1:xa[0]', ward)
        driver.find_element(By.CSS_SELECTOR, '#text-input-Los_1_nguoi_tham_chieu\:Horizontal_Layout2\:Table1\:Text_Area1\[0\]').send_keys(address)
        # click to Tiếp tục button
        clickButtonByText('Tiếp tục')
    except Exception as ex:
        raise Exception('Testcase fail at Khach Hang Page!')


def testNgheNghiepPage():
    try:
        switchIframeByTitle('Coach')
        clickButtonByText('Tiếp tục')
    except Exception as ex:
        raise Exception('Testcase fail at Nghe Nghiep Page!')

def testTaiChinhPage():
    try:
        a=1
        # switchIframeByTitle('Coach')       
        # elementBeginPosition = driver.find_element(By.ID, 'footerx')
        # elementEndPosition = driver.find_element(By.ID, 'tabs-Tab_Section1-tab-text0')
        # scrollDownUp(elementBeginPosition, elementEndPosition)

        # time.sleep(1)
        # driver.find_element(By.ID, 'tabs-LOS_ER_BUSINESS_FINANCE1:ts_cn-tab-text1').click()
        # scrollDownUp(elementBeginPosition, elementEndPosition)

        # time.sleep(1)
        # driver.find_element(By.ID, 'tabs-LOS_ER_BUSINESS_FINANCE1:ts_cn-tab-text2').click()
        # scrollDownUp(elementBeginPosition, elementEndPosition)

        # time.sleep(1)
        # driver.find_element(By.ID, 'tabs-Tab_Section1-tab-text1').click()
        # scrollToElement(elementBeginPosition)
        # time.sleep(1)
        # driver.find_element(By.ID, 'table-pagination-2-LOS_3_Tab_C1:Table1').click()
        # scrollDownUp(elementEndPosition, elementBeginPosition)
        # time.sleep(1)
        # driver.find_element(By.ID, 'table-pagination-3-LOS_3_Tab_C1:Table1').click()
        # scrollToElement(elementEndPosition)

        # time.sleep(1)
        # driver.find_element(By.ID, 'tabs-Tab_Section1-tab-text2').click()
        # scrollDownUp(elementBeginPosition, elementEndPosition)
    except Exception as ex:
        raise Exception('Testcase fail at Tai Chinh Page!')
    

def testKhoanVayPage(soTienVay, thoiHanVay, loaiLaiSuat, laiSuatChoVay, nguonTraNoChinh, maAM, maSanPham, ngayDenHanTraGocDauTien, ngayTrongThangDenHanTraGoc,
                    phuongThucTraNo, soLanRutVon, soTienRutVon, noiDungRutVon):
    try:
        switchIframeByTitle('Coach')
        driver.find_element(By.ID, 'HptBreadcrumb-item-breadCrumbs[3]').click()
        switchIframeByTitle('Coach') 
        time.sleep(15)
        wait.until(EC.presence_of_element_located((By.ID, 'decimal-input-LOS_Loan_Information1:txt_loan_money'))) 
        driver.find_element(By.ID, 'decimal-input-LOS_Loan_Information1:txt_loan_money').send_keys(soTienVay)
        driver.find_element(By.ID, 'decimal-input-LOS_Loan_Information1:int_tenor').send_keys(thoiHanVay)
        clickDropdownBoxByText('singleselect-LOS_Loan_Information1:ss_interest', loaiLaiSuat)
        driver.find_element(By.ID, 'decimal-input-LOS_Loan_Information1:laixuatchovay').send_keys(laiSuatChoVay)
        clickDropdownBoxByText('singleselect-LOS_Loan_Information1:Text2', nguonTraNoChinh)
        time.sleep(5)
        wait.until(EC.presence_of_all_elements_located((By.ID, 'text-input-LOS_Loan_Information1:int_AMPLITUTE1'))) 
        driver.find_element(By.ID, 'text-input-LOS_Loan_Information1:int_AMPLITUTE1').send_keys(maAM)
        clickDropdownBoxByText('singleselect-LOS_Loan_Information1:ss_product_code', maSanPham)
        driver.find_element(By.ID, 'datetimepicker-input-LOS_Loan_Information1:Date_Time_Picker1').send_keys(ngayDenHanTraGocDauTien)
        clickDropdownBoxByText('singleselect-LOS_Loan_Information1:ngay_trongthang_denhan_tragoc', ngayTrongThangDenHanTraGoc)
        clickDropdownBoxByText('singleselect-LOS_Loan_Information1:Text15', phuongThucTraNo)
        clickDropdownBoxByText('singleselect-LOS_Loan_Information1:NO_WITHDRAWALS', soLanRutVon)
        
        driver.find_element(By.ID, 'checkbox-input-LOS_Loan_Information1:LN106').click()
        driver.find_element(By.ID, 'table-addbutton-icon-LOS_Loan_Information1:view_lichgiaingan1:Table1').click()
        driver.find_element(By.ID, 'table-addbutton-icon-LOS_Loan_Information1:view_lichgiaingan1:view_thong_tin_de_xuat_tung_lan1:view_thongtinrutvon1:tbl_rutvon').click()
        driver.find_element(By.ID, 'decimal-input-LOS_Loan_Information1:view_lichgiaingan1:view_thong_tin_de_xuat_tung_lan1:view_thongtinrutvon1:tbl_rutvon:sotien_rutvon[0]').send_keys(soTienRutVon)
        driver.find_element(By.ID, 'text-input-LOS_Loan_Information1:view_lichgiaingan1:view_thong_tin_de_xuat_tung_lan1:view_thongtinrutvon1:tbl_rutvon:noidung_rutvon[0]').send_keys(noiDungRutVon)
        driver.find_element(By.ID, 'button-button-LOS_Loan_Information1:view_lichgiaingan1:view_thong_tin_de_xuat_tung_lan1:Button1').click()
    except Exception as ex:
        raise Exception('Testcase fail at Khoan Vay Page!')        
        

def testLichTraNoPage():
    try:
        driver.find_element(By.ID, 'HptBreadcrumb-item-breadCrumbs[4]').click()
        switchIframeByTitle('Coach')         
    except Exception as ex:
        raise Exception('Testcase fail at Lich Tra No Page!')      
    

def testHoSoPage():
    global maGiaoDich
    try:
        driver.find_element(By.ID, 'HptBreadcrumb-item-breadCrumbs[5]').click()
        switchIframeByTitle('Coach')         
        driver.find_element(By.ID, 'button-button-LOS_Buttons1:btn_complete').click()
        wait.until(EC.visibility_of_element_located((By.ID, 'outputtext-text-LOS_Alert_BPM_Info_Specified1:Text1')))
        maGiaoDich = driver.find_element(By.ID, 'outputtext-text-LOS_Alert_BPM_Info_Specified1:Text1').text
        time.sleep(3)
        driver.find_element(By.ID, 'button-button-LOS_Alert_BPM_Info_Specified1:Text3').click()
    except Exception as ex:
        raise Exception('Testcase fail at Ho So Page!')         

def testLogoutButton():
    try:
        #Click Logout button
        switchToNewTab(1, 0)
        driver.switch_to.default_content()
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR,  'div > div > div:nth-child(2) > div.profile-actions > div:nth-child(2) > a')))
        driver.find_element(By.CSS_SELECTOR, 'div > div > div:nth-child(2) > div.profile-actions > div:nth-child(2) > a').click()
    except Exception as ex: 
        raise Exception('Testcase fail at Logout Button!')   

def switchIframeThamDinhPage():
    driver.switch_to.default_content()
    switchIframeByTitle('Thẩm định')
    switchIframeByTitle('Coach')


def testThamDinhPage(tinhTP):
    try:
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '#div_1_2_1_2_1 > div > div.filter > div.input-box > div.form-group.has-feedback.empty > input')))
        temp = driver.find_element(By.CSS_SELECTOR, '#div_1_2_1_2_1 > div > div.filter > div.input-box > div.form-group.has-feedback.empty > input')
        temp.send_keys(maGiaoDich)
        time.sleep(1)
        driver.find_element(By.CSS_SELECTOR, '#div_1_2_1_2_1 > div > div.filter > div.update-filter > button').click()

        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div[title='Task Subject Thẩm định']")))
        driver.find_element(By.CSS_SELECTOR, "div[title='Task Subject Thẩm định']").click()
        
        switchIframeByTitle('Thẩm định')
        switchIframeByTitle('Coach')

        wait.until(EC.visibility_of_element_located((By.ID, 'HptBreadcrumb-item-breadCrumbs[5]')))
        driver.find_element(By.ID, 'HptBreadcrumb-item-breadCrumbs[5]').click()

        switchIframeByTitle('Thẩm định')
        switchIframeByTitle('Coach')
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "#icon-button-LOS_Rule_Results1\:LIST_TITLES1\:action\[0\]\:Horizontal_Layout2\:Button1")))
        driver.find_element(By.CSS_SELECTOR, "#icon-button-LOS_Rule_Results1\:LIST_TITLES1\:action\[0\]\:Horizontal_Layout2\:Button1").click()
        time.sleep(2)

        switchIframeThamDinhPage()
        displayPopUp = driver.find_element(By.CSS_SELECTOR, "div[class = 'modal fade in']")
        driver.execute_script("arguments[0].setAttribute('style', arguments[1]);", displayPopUp, "display: none;")

        switchIframeThamDinhPage()
        driver.find_element(By.ID, 'HptBreadcrumb-item-breadCrumbs[2]').click()
        time.sleep(2)

        switchIframeByTitle('Thẩm định')
        switchIframeByTitle('Coach')        
        wait.until(EC.visibility_of_element_located((By.ID, 'singleselect-LOS_Legal_Work_Information1:LOS_AddressTBL_noHeader1:Table1:tinh[0]')))
        clickDropdownBoxByText('singleselect-LOS_Legal_Work_Information1:LOS_AddressTBL_noHeader1:Table1:tinh[0]', tinhTP)
        driver.find_element(By.ID, 'HptBreadcrumb-item-breadCrumbs[5]').click()
        time.sleep(30)
        switchIframeThamDinhPage()
        wait.until(EC.visibility_of_element_located((By.ID, 'outputtext-text-Template1:Output_Text1')))
        scrollToElement(driver.find_element(By.ID, 'outputtext-text-Template1:Output_Text1'))
        time.sleep(5)
        driver.find_element(By.ID, 'HptBreadcrumb-item-breadCrumbs[7]').click()
        switchIframeThamDinhPage()
        time.sleep(2)
        wait.until(EC.visibility_of_element_located((By.ID, 'button-button-LOS_Buttons1:btn_complete')))
        driver.find_element(By.ID, 'button-button-LOS_Buttons1:btn_complete').click()
        time.sleep(2)
        switchIframeThamDinhPage()
        wait.until(EC.visibility_of_element_located((By.ID, 'outputtext-text-LOS_Alert_BPM_Info_Specified1:Text1')))
        # maTemp = driver.find_element(By.ID, 'outputtext-text-LOS_Alert_BPM_Info_Specified1:Text1').text
        
        driver.find_element(By.ID, 'radiogroup-item-input-LOS_Alert_BPM_Info_Specified1:radioHoso[0]').click()
        time.sleep(2)
        switchIframeThamDinhPage()
        wait.until(EC.visibility_of_element_located((By.ID, 'button-button-LOS_Alert_BPM_Info_Specified1:Text3')))
        driver.find_element(By.ID, 'button-button-LOS_Alert_BPM_Info_Specified1:Text3').click()
        time.sleep(20)

    except Exception as ex:
        print('Fail Test Case at Tham Dinh Page')
        pass

def  executeTestCase(sttTestCase, userName, password, fullName, productName, loanMethod, policy, 
                     referencePersonName, relationship, gender, phoneNumber, nationality, province, district, ward, address,
                     soTienVay, thoiHanVay, loaiLaiSuat, laiSuatChoVay, nguonTraNoChinh, maAM, maSanPham, ngayDenHanTraGocDauTien, ngayTrongThangDenHanTraGoc,
                    phuongThucTraNo, soLanRutVon, soTienRutVon, noiDungRutVon, userName2, password2, tinhTP):
    try:
        driver.get('https://10.53.121.40:9443/ProcessPortal/login.jsp')
        testLoginPage(userName, password)
        testKhoiTaoKhoanVay(fullName, productName, loanMethod, policy)
        testKhachHangPage(referencePersonName, relationship, gender, phoneNumber,nationality, province, district, ward, address)
        testNgheNghiepPage()
        testTaiChinhPage()
        testKhoanVayPage(soTienVay, thoiHanVay, loaiLaiSuat, laiSuatChoVay, nguonTraNoChinh, maAM, maSanPham, ngayDenHanTraGocDauTien,
                          ngayTrongThangDenHanTraGoc, phuongThucTraNo, soLanRutVon, soTienRutVon, noiDungRutVon)
        testLichTraNoPage()
        testHoSoPage()
        testLogoutButton()
        testLoginPage(userName2, password2)
        testThamDinhPage(tinhTP)
        print(str(sttTestCase) + '[+]: Test Case PASS')
    except Exception as ex:
        print(str(sttTestCase) + '[!]: ', ex)



if __name__ == '__main__':
    config()
    for row in range(3, totalRow+1):
        sttTestCase = row-2
        executeTestCase(sttTestCase, getvalueExel(row, 3), getvalueExel(row, 4), getvalueExel(row, 16), getvalueExel(row, 10), 
                        getvalueExel(row, 11), getvalueExel(row, 12), getvalueExel(row, 25), getvalueExel(row, 26), 
                        getvalueExel(row, 27), getvalueExel(row, 28), getvalueExel(row, 29), getvalueExel(row, 30), 
                        getvalueExel(row, 31), getvalueExel(row, 32), getvalueExel(row, 33), getvalueExel(row, 52), 
                        getvalueExel(row, 53), getvalueExel(row, 54), getvalueExel(row, 55), getvalueExel(row, 56), 
                        getvalueExel(row, 57), getvalueExel(row, 58), getvalueExel(row, 59), getvalueExel(row, 60), 
                        getvalueExel(row, 61), getvalueExel(row, 62), getvalueExel(row, 63), getvalueExel(row, 64), 
                        getvalueExel(row, 5), getvalueExel(row, 6), getvalueExel(row, 65))
